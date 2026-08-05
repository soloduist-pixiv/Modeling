#!/usr/bin/env python3
"""Export mooring results (metrics.json) to a multi-sheet Excel workbook."""
from __future__ import annotations

import json
import os
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), "..", "results")
METRICS = os.path.join(ROOT, "metrics.json")
OUT_XLSX = os.path.join(ROOT, "results_tables.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _flatten_pipes(row: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(row)
    pipes = out.pop("pipe_angles_deg", None)
    if isinstance(pipes, list):
        for i, a in enumerate(pipes, 1):
            out[f"pipe{i}_deg"] = a
    return out


def _write_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]], preferred_cols: Optional[List[str]] = None):
    ws = wb.create_sheet(title)
    if not rows:
        ws["A1"] = "(无数据)"
        return
    flat = [_flatten_pipes(r) for r in rows]
    cols: List[str] = []
    if preferred_cols:
        for c in preferred_cols:
            if any(c in r for r in flat):
                cols.append(c)
    for r in flat:
        for k in r.keys():
            if k not in cols:
                cols.append(k)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN
    for i, r in enumerate(flat, 2):
        for j, c in enumerate(cols, 1):
            val = r.get(c, "")
            if isinstance(val, float):
                val = round(val, 6)
            cell = ws.cell(i, j, val)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
    for j, c in enumerate(cols, 1):
        width = min(max(10, len(str(c)) + 2), 28)
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _summary_rows(report: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = [
        {"项目": "Freeze/修订", "数值": report.get("revision", "")},
        {"项目": "软裕度-钢桶倾角(°)", "数值": report.get("soft_limits", {}).get("barrel_deg")},
        {"项目": "软裕度-锚端夹角(°)", "数值": report.get("soft_limits", {}).get("anchor_deg")},
        {"项目": "软裕度-吃水(m)", "数值": report.get("soft_limits", {}).get("draft_m")},
    ]
    q2 = report.get("q2_ball_search_A", {})
    if q2.get("found"):
        b = q2["best"]
        rows.extend(
            [
                {"项目": "Q2最小球重A(kg)", "数值": b.get("ball_mass")},
                {"项目": "Q2钢桶倾角(°)", "数值": b.get("barrel_angle_deg")},
                {"项目": "Q2锚端夹角(°)", "数值": b.get("anchor_angle_deg")},
                {"项目": "Q2游动半径(m)", "数值": b.get("swim_radius_m")},
                {"项目": "Q2吃水(m)", "数值": b.get("draft_m")},
            ]
        )
    q3 = report.get("q3_design", {})
    best = q3.get("best") or {}
    rows.extend(
        [
            {"项目": "Q3推荐型号", "数值": best.get("chain_type")},
            {"项目": "Q3推荐链长(m)", "数值": best.get("chain_length")},
            {"项目": "Q3推荐节数", "数值": best.get("n_links")},
            {"项目": "Q3推荐球重(kg)", "数值": best.get("ball_mass")},
            {"项目": "Q3训练可行数", "数值": q3.get("n_feasible_train")},
            {"项目": "Q3说明", "数值": q3.get("note")},
            {
                "项目": "稠密扫描硬约束失败数",
                "数值": (q3.get("dense_scan") or {}).get("n_fail_hard"),
            },
        ]
    )
    return rows


def export_xlsx(metrics_path: str = METRICS, out_path: str = OUT_XLSX) -> str:
    with open(metrics_path, encoding="utf-8") as f:
        report = json.load(f)

    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    _write_sheet(wb, "00_摘要", _summary_rows(report), ["项目", "数值"])

    state_cols = [
        "family",
        "v_wind",
        "v_cur",
        "depth",
        "ball_mass",
        "chain_type",
        "chain_length",
        "n_links",
        "draft_m",
        "swim_radius_m",
        "barrel_angle_deg",
        "anchor_angle_deg",
        "pipe1_deg",
        "pipe2_deg",
        "pipe3_deg",
        "pipe4_deg",
        "grounded_length_m",
        "suspended_chain_m",
        "H_N",
        "H_buoy_N",
        "vertical_closure_err_m",
        "constraint_barrel_ok",
        "constraint_anchor_ok",
        "margin_barrel_ok",
        "margin_anchor_ok",
        "margin_draft_ok",
        "ok",
        "message",
    ]
    _write_sheet(wb, "01_Q1Q2固定球重", report.get("q1_q2_fixed_ball") or [], state_cols)

    q2a = report.get("q2_ball_search_A") or {}
    q2b = report.get("q2_ball_search_B") or {}
    q2_rows = []
    if q2a.get("best"):
        r = dict(q2a["best"])
        r["tag"] = "A_min_mass"
        q2_rows.append(r)
    if q2a.get("best_by_swim"):
        r = dict(q2a["best_by_swim"])
        r["tag"] = "A_min_swim"
        q2_rows.append(r)
    if q2b.get("best"):
        r = dict(q2b["best"])
        r["tag"] = "B_min_mass"
        q2_rows.append(r)
    _write_sheet(wb, "02_Q2调球重", q2_rows, ["tag"] + state_cols)

    _write_sheet(wb, "03_族对照AB", report.get("family_contrast") or [], state_cols)

    q3 = report.get("q3_design") or {}
    _write_sheet(wb, "04_Q3训练情景", q3.get("best_train_rows") or [], state_cols)

    hold = q3.get("holdout") or {}
    _write_sheet(wb, "05_Q3_holdout", hold.get("rows") or [], state_cols)

    design_cols = [
        "chain_type",
        "chain_length",
        "n_links",
        "ball_mass",
        "ok_all",
        "margin_all",
        "worst_barrel",
        "worst_anchor",
        "max_draft",
        "mean_swim",
        "mean_draft",
        "utility",
    ]
    design_rows = []
    if q3.get("best"):
        r = dict(q3["best"])
        r["tag"] = "recommended"
        design_rows.append(r)
    if q3.get("alt_min_swim"):
        r = dict(q3["alt_min_swim"])
        r["tag"] = "alt_min_swim"
        design_rows.append(r)
    if q3.get("alt_min_draft"):
        r = dict(q3["alt_min_draft"])
        r["tag"] = "alt_min_draft"
        design_rows.append(r)
    for r in q3.get("top_margin") or []:
        rr = dict(r)
        rr["tag"] = "top_margin"
        design_rows.append(rr)
    for r in q3.get("top_hard") or []:
        rr = dict(r)
        rr["tag"] = "top_hard"
        design_rows.append(rr)
    _write_sheet(wb, "06_Q3推荐与备选", design_rows, ["tag"] + design_cols)
    _write_sheet(wb, "07_Q3_Pareto", q3.get("pareto_front") or [], design_cols)

    dense = q3.get("dense_scan") or {}
    _write_sheet(wb, "08_稠密扫描最劣", dense.get("worst_rows") or [], state_cols)

    abl = report.get("ablation_ball_buoyancy") or []
    _write_sheet(wb, "09_球浮力消融", abl, ["ablation"] + state_cols)
    sens = report.get("sensitivity_rho") or []
    _write_sheet(wb, "10_密度敏感性", sens, ["rho"] + state_cols)

    # Q2 sweep table for quick plotting in Excel
    sweep_rows = []
    try:
        import sys

        sys.path.insert(0, os.path.dirname(__file__))
        from mooring_solve import solve_static, result_to_public

        for m in range(1200, 4001, 100):
            pub = result_to_public(solve_static(36.0, 18.0, float(m), "II", 22.05, family="A"))
            sweep_rows.append(pub)
    except Exception as e:
        sweep_rows = [{"ok": False, "message": f"sweep failed: {e}"}]
    _write_sheet(wb, "11_Q2球重扫描", sweep_rows, state_cols)

    # integer link examples
    ex = report.get("integer_link_examples") or {}
    ex_rows = [{"name": k, "length_m": v[0], "n_links": v[1]} for k, v in ex.items()]
    _write_sheet(wb, "12_整数链节示例", ex_rows, ["name", "length_m", "n_links"])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    path = export_xlsx()
    print("Wrote", path)


if __name__ == "__main__":
    main()
