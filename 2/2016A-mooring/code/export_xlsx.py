#!/usr/bin/env python3
"""Export mooring results (metrics.json) to a multi-sheet Excel workbook."""
from __future__ import annotations

import json
import os
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), "..", "results")
METRICS = os.path.join(ROOT, "metrics.json")
OUT_XLSX = os.path.join(ROOT, "results_tables.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _flatten_pipes(row: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(row)
    pipes = out.pop("pipe_angles_deg", None)
    if isinstance(pipes, list):
        for i, a in enumerate(pipes, 1):
            out[f"pipe{i}_deg"] = a
    res = out.pop("residual", None)
    if isinstance(res, dict):
        out["res_H_rel"] = res.get("horizontal_rel")
        out["res_V_N"] = res.get("vertical_abs_N")
        out["res_geom_m"] = res.get("geometry_abs_m")
    for k, v in list(out.items()):
        if isinstance(v, (list, dict)):
            out[k] = json.dumps(v, ensure_ascii=False)
    return out


def _write_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]], preferred_cols: Optional[List[str]] = None):
    ws = wb.create_sheet(title)
    if not rows:
        ws["A1"] = "(无数据)"
        return
    flat = [_flatten_pipes(r) for r in rows]
    cols: List[str] = []
    if preferred_cols:
        for c in preferred_cols:
            if any(c in r for r in flat):
                cols.append(c)
    for r in flat:
        for k in r.keys():
            if k not in cols:
                cols.append(k)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN
    for i, r in enumerate(flat, 2):
        for j, c in enumerate(cols, 1):
            val = r.get(c, "")
            if isinstance(val, float):
                val = round(val, 6)
            cell = ws.cell(i, j, val)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
    for j, c in enumerate(cols, 1):
        width = min(max(10, len(str(c)) + 2), 28)
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _summary_rows(report: Dict[str, Any]) -> List[Dict[str, Any]]:
    lim = report.get("hard_limits", {})
    rows = [
        {"项目": "修订", "数值": report.get("revision", "")},
        {"项目": "硬约束-钢桶倾角(°)", "数值": lim.get("barrel_deg")},
        {"项目": "硬约束-锚端夹角(°)", "数值": lim.get("anchor_deg")},
    ]
    q2 = report.get("q2_ball_search_A", {})
    if q2.get("found"):
        b = q2["best"]
        rows.extend([
            {"项目": "Q2最小球重A(kg)", "数值": b.get("ball_mass")},
            {"项目": "Q2钢桶倾角(°)", "数值": b.get("barrel_angle_deg")},
            {"项目": "Q2锚端夹角(°)", "数值": b.get("anchor_angle_deg")},
            {"项目": "Q2游动半径(m)", "数值": b.get("swim_radius_m")},
            {"项目": "Q2吃水(m)", "数值": b.get("draft_m")},
            {"项目": "Q2紧约束", "数值": q2.get("binding")},
        ])
    q3 = report.get("q3_design", {})
    best = q3.get("best") or {}
    ver = q3.get("verification") or {}
    rows.extend([
        {"项目": "Q3推荐型号", "数值": best.get("chain_type")},
        {"项目": "Q3推荐链长(m)", "数值": best.get("chain_length")},
        {"项目": "Q3推荐节数", "数值": best.get("n_links")},
        {"项目": "Q3推荐球重(kg)", "数值": best.get("ball_mass")},
        {"项目": "Q3最劣钢桶倾角(°)", "数值": best.get("worst_barrel")},
        {"项目": "Q3最劣锚端夹角(°)", "数值": best.get("worst_anchor")},
        {"项目": "Q3最大吃水(m)", "数值": best.get("max_draft")},
        {"项目": "Q3最小干舷(m)", "数值": best.get("min_freeboard")},
        {"项目": "Q3权重随机化胜出率", "数值": best.get("win_rate")},
        {"项目": "可行设计数", "数值": q3.get("n_feasible_designs")},
        {"项目": "Pareto 规模", "数值": q3.get("pareto_size")},
        {"项目": "稠密扫描点数", "数值": ver.get("dense_n")},
        {"项目": "稠密扫描硬约束失败数", "数值": ver.get("dense_hard_fail")},
        {"项目": "参数扰动下约束满足率", "数值": (ver.get("reliability") or {}).get("satisfaction_rate")},
        {"项目": "解析吃水下界(36m/s,1.5m/s)", "数值": (q3.get("analytic_draft_bound") or {}).get("vw36_vc1.5")},
    ])
    return rows


def export_xlsx(metrics_path: str = METRICS, out_path: str = OUT_XLSX) -> str:
    with open(metrics_path, encoding="utf-8") as f:
        report = json.load(f)

    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    _write_sheet(wb, "00_摘要", _summary_rows(report), ["项目", "数值"])

    state_cols = [
        "family",
        "v_wind",
        "v_cur",
        "depth",
        "ball_mass",
        "chain_type",
        "chain_length",
        "n_links",
        "draft_m",
        "swim_radius_m",
        "barrel_angle_deg",
        "anchor_angle_deg",
        "pipe1_deg",
        "pipe2_deg",
        "pipe3_deg",
        "pipe4_deg",
        "grounded_length_m",
        "suspended_chain_m",
        "H_N",
        "H_buoy_N",
        "ball_drag_N",
        "freeboard_m",
        "vertical_closure_err_m",
        "res_H_rel",
        "res_V_N",
        "res_geom_m",
        "constraint_barrel_ok",
        "constraint_anchor_ok",
        "ok",
        "message",
    ]
    _write_sheet(wb, "01_Q1Q2固定球重", report.get("q1_q2_fixed_ball") or [], state_cols)

    q2a = report.get("q2_ball_search_A") or {}
    q2b = report.get("q2_ball_search_B") or {}
    q2_rows = []
    if q2a.get("best"):
        r = dict(q2a["best"])
        r["tag"] = "A_min_mass"
        q2_rows.append(r)
    if q2b.get("best"):
        r = dict(q2b["best"])
        r["tag"] = "B_min_mass"
        q2_rows.append(r)
    _write_sheet(wb, "02_Q2调球重", q2_rows, ["tag"] + state_cols)
    _write_sheet(wb, "03_Q2球重扫描", report.get("q2_ball_sweep") or [], state_cols)
    _write_sheet(wb, "04_族对照AB", report.get("family_contrast") or [], state_cols)

    q3 = report.get("q3_design") or {}
    _write_sheet(wb, "05_Q3最劣情景", q3.get("best_worst_case_rows") or [], ["role"] + state_cols)

    ver = q3.get("verification") or {}
    _write_sheet(wb, "06_Q3典型情景", ver.get("typical_rows") or [], state_cols)

    design_cols = [
        "chain_type", "chain_length", "n_links", "ball_mass", "m_min_continuous",
        "worst_barrel", "worst_anchor", "max_draft", "max_swim", "min_freeboard",
        "topsis", "win_rate", "top3_rate",
    ]
    design_rows = []
    for tag, key in [("recommended", "best"), ("alt_min_draft", "alt_min_draft"), ("alt_min_swim", "alt_min_swim")]:
        if q3.get(key):
            r = dict(q3[key])
            r["tag"] = tag
            design_rows.append(r)
    _write_sheet(wb, "07_Q3推荐与备选", design_rows, ["tag"] + design_cols)
    _write_sheet(wb, "08_Q3_Pareto排序", q3.get("ranked_pareto") or [], design_cols)
    _write_sheet(wb, "09_各型号最小球重", q3.get("minimal_ball_table") or [], design_cols)

    bound = q3.get("analytic_draft_bound") or {}
    _write_sheet(wb, "10_解析吃水下界", [{"情景": k, "吃水下界_m": v} for k, v in bound.items()],
                 ["情景", "吃水下界_m"])
    _write_sheet(wb, "11_吃水下界曲线", report.get("draft_bound_curve") or [],
                 ["v_cur", "v_wind", "draft_min_m"])

    mono = report.get("monotonicity") or {}
    mono_rows = [{"axis": axis, **{k: v for k, v in d.items() if k != "rows"}} for axis, d in mono.items()]
    _write_sheet(wb, "12_单调性判定", mono_rows, ["axis", "barrel_trend", "anchor_trend", "draft_trend", "swim_trend"])

    _write_sheet(wb, "13_球浮力消融", report.get("ablation_ball_buoyancy") or [], ["ablation"] + state_cols)
    _write_sheet(wb, "14_球阻力消融", report.get("ablation_ball_drag") or [], ["ablation"] + state_cols)
    _write_sheet(wb, "15_密度敏感性", report.get("sensitivity_rho") or [], ["rho"] + state_cols)
    _write_sheet(wb, "16_链阻力敏感性", report.get("sensitivity_chain_drag") or [],
                 ["chain_drag_scale"] + state_cols)
    _write_sheet(wb, "17_离散收敛", report.get("grid_refinement") or [],
                 ["subdivision", "n_elements", "anchor_angle_deg", "span_x_m", "span_y_m"])

    rel = ver.get("reliability") or {}
    _write_sheet(wb, "18_参数不确定性", [{"指标": k, "数值": v} for k, v in rel.items()], ["指标", "数值"])
    rbd = ver.get("reliability_based") or {}
    _write_sheet(wb, "19_可靠性球重曲线", rbd.get("rate_profile") or [], ["ball_mass", "rate"])

    relaxed = report.get("q3_relaxed") or {}
    relaxed_cols = ["chain_type", "chain_length", "n_links", "ball_mass", "worst_barrel_relaxed",
                    "worst_barrel_full_box", "worst_anchor", "max_draft", "max_swim", "min_freeboard"]
    _write_sheet(wb, "20_干舷优先方案", relaxed.get("table") or [], relaxed_cols)
    strat = []
    if q3.get("best"):
        b = dict(q3["best"]); b["策略"] = "S1 严格 5° 全包络"; strat.append(b)
    if relaxed.get("best"):
        b = dict(relaxed["best"]); b["策略"] = "S2 干舷优先"; strat.append(b)
    _write_sheet(wb, "21_双策略对比", strat, ["策略"] + relaxed_cols)

    ex = report.get("integer_link_examples") or {}
    _write_sheet(wb, "22_整数链节示例",
                 [{"name": k, "length_m": v[0], "n_links": v[1]} for k, v in ex.items()],
                 ["name", "length_m", "n_links"])
    _write_sheet(wb, "23_链条投影宽度",
                 [{"型号": t, **d} for t, d in (report.get("chain_drag_widths") or {}).items()],
                 ["型号", "volume_equivalent", "link_geometry"])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    path = export_xlsx()
    print("Wrote", path)


if __name__ == "__main__":
    main()
